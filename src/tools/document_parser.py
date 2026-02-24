"""
Document parser utility — extracts text from various file formats.

NOT a tool — a helper function used by main.py to parse documents
sent via Telegram before passing them to the agent.

Supported formats:
- Text files: .txt, .md, .csv, .json, .log, .py, .js, .ts, .html, .xml, .yaml, .toml, .ini, .cfg
- PDF: .pdf (via pypdf)
- Word: .docx (via python-docx)
- Excel: .xlsx (via openpyxl)

Unsupported: .doc, .xls (legacy formats requiring LibreOffice)
"""

from __future__ import annotations

import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# Max characters to return (prevents context overload, ~8K tokens)
MAX_CONTENT_LENGTH = 30_000

# Text-based file extensions (read as UTF-8)
_TEXT_EXTENSIONS = {
    ".txt", ".md", ".csv", ".json", ".log",
    ".py", ".js", ".ts", ".jsx", ".tsx",
    ".html", ".htm", ".xml", ".svg",
    ".yaml", ".yml", ".toml", ".ini", ".cfg",
    ".sh", ".bat", ".ps1",
    ".sql", ".env", ".gitignore",
    ".c", ".cpp", ".h", ".java", ".go", ".rs", ".rb",
}


def _truncate(text: str) -> str:
    """Truncate text to MAX_CONTENT_LENGTH with a notice."""
    if len(text) <= MAX_CONTENT_LENGTH:
        return text
    return text[:MAX_CONTENT_LENGTH] + "\n\n[... обрезано, файл слишком большой]"


def parse_document(file_path: Path, file_name: str) -> str:
    """Extract text content from a document file.

    Args:
        file_path: Path to the downloaded file on disk.
        file_name: Original file name (used for extension detection).

    Returns:
        Extracted text content (truncated if too long).
    """
    ext = Path(file_name).suffix.lower()

    try:
        # Plain text files
        if ext in _TEXT_EXTENSIONS:
            return _parse_text(file_path)

        # PDF
        if ext == ".pdf":
            return _parse_pdf(file_path)

        # Word (.docx)
        if ext == ".docx":
            return _parse_docx(file_path)

        # Excel (.xlsx)
        if ext == ".xlsx":
            return _parse_xlsx(file_path)

        # Unsupported
        return f"[Формат {ext} не поддерживается для чтения содержимого]"

    except Exception as exc:
        logger.error("Failed to parse document %s: %s", file_name, exc)
        return f"[Ошибка при чтении файла: {exc}]"


def _parse_text(file_path: Path) -> str:
    """Read a text file with encoding fallback."""
    try:
        text = file_path.read_text(encoding="utf-8", errors="replace")
    except UnicodeDecodeError:
        text = file_path.read_text(encoding="cp1251", errors="replace")
    return _truncate(text)


def _parse_pdf(file_path: Path) -> str:
    """Extract text from a PDF file using pypdf."""
    from pypdf import PdfReader

    reader = PdfReader(str(file_path))
    pages = []
    for page in reader.pages:
        text = page.extract_text() or ""
        if text.strip():
            pages.append(text)

    if not pages:
        return "[PDF файл не содержит текста (возможно, отсканированный)]"

    return _truncate("\n\n".join(pages))


def _parse_docx(file_path: Path) -> str:
    """Extract text from a .docx file using python-docx."""
    from docx import Document

    doc = Document(str(file_path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]

    if not paragraphs:
        return "[Документ Word пуст]"

    return _truncate("\n\n".join(paragraphs))


def _parse_xlsx(file_path: Path) -> str:
    """Extract data from an .xlsx file as a text table."""
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    sheets = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):  # skip empty rows
                rows.append(" | ".join(cells))

        if rows:
            header = f"--- Лист: {sheet_name} ---"
            sheets.append(f"{header}\n{chr(10).join(rows)}")

    wb.close()

    if not sheets:
        return "[Excel файл пуст]"

    return _truncate("\n\n".join(sheets))
